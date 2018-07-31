from django.shortcuts import render, get_object_or_404, redirect
from django.http import Http404, HttpResponse
from .models import Product
from .forms import ProductForm


def index(request):
    products = Product.objects.all()
    context = {'products': products}
    return render(request, 'inventory/index.html', context)


def detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'inventory/detail.html', {'product': product})


def addnew(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        print(dir(form))
        if form.is_valid():
            # product = form.save(commit=True)
            form.save()

            # product = Product()
            # product.name = form.cleaned_data['name']
            # product.cetagory = form.cleaned_data['cetagory']
            # product.supplier = form.cleaned_data['supplier']
            # product.unit_price = form.cleaned_data['unit_price']
            # product.description = form.cleaned_data['description']
            # product.save()
            # return redirect('detail', pk=product.pk)
            return redirect('index')
    else:
        form = ProductForm()
    return render(request, 'inventory/new.html', {'form': form})


def edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/edit.html', {'form': form})


#
'''
def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.title,
            obj.description,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"
'''

def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Title"),
        smart_str(u"Description"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.title),
            smart_str(obj.description),
        ])
    return response
export_csv.short_description = u"Export CSV"